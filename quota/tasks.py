from background_task import background
from quota.models import UploadQuota, Quota
from openpyxl import load_workbook
from io import BytesIO

@background
def process_quota(quota_id):
    quota = UploadQuota.objects.get(id=quota_id)
    file = quota.quotafile
    wb = load_workbook(filename=BytesIO(file))
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
        Quota.objects.create(
            user = row[0],
            type = row[1],
            amount = row[2]
        )
